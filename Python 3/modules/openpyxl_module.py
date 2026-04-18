#!/usr/bin/env python3

# PRE: pip3 install openpyxl

import openpyxl

# Convert between Excel column letters and integers.
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('spreadsheet.xlsx')

print(wb.sheetnames)

sheet = wb['Numbers']
sheet['A1':'B2'] # Rectangular slice.

print(sheet.title)
print(sheet.max_row)
print(sheet.max_column)

cell = sheet['A1']
value = cell.value
row = cell.row
column = cell.column # Integer value, not letter(s).
print(cell.coordinate) # A1

print(row)
print(column)
print(value)

wb = openpyxl.Workbook() # Create a new blank workbook.
wb.create_sheet(index=0, title='New Sheet') # Makes the new sheet the active sheet.
sheet = wb.active
sheet['A1'] = 'Hello, Excel!' # Write to a cell in a worksheet.
wb.save('test.xlsx') # Overwrites existing.


